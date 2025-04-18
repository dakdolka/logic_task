import os
from fastapi import HTTPException
from sqlalchemy import text, insert, select, func, cast, Integer, and_, update
# from db.models import metadata_obj
from sqlalchemy.orm import aliased
from sqlalchemy.orm import joinedload, selectinload, contains_eager
from data.database import Base, async_engine, async_session_factory
from data.models import Info, OrdderConstructor, Order, pack
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


class Orm:
    @staticmethod
    async def create_all():
        async with async_engine.begin() as conn:
            async_engine.echo = False
            # await conn.run_sync(Base.metadata.drop_all)
            await conn.run_sync(Base.metadata.create_all)
            print('tables created')
            
            async_engine.echo = True
    
    @staticmethod
    async def insert_or_upd_info(info):
        async with async_session_factory() as session:
            print('upd_info')
            async_engine.echo = False
            info = info[['Наименование', 'Артикул', 'Упаковка высота (мм)', "Упаковка длина (см)", "Упаковка ширина (см)", "УПАКОВКА FBO (Общие)"]]
            data = info.values.tolist()
            dop = []
            for i in range(len(data)):
                # print(data[i])
                # exit()
                if not pd.isna(data[i][1]):
                    data[i] = {
                        'id': data[i][1] ,
                        'name': data[i][0] if not pd.isna(data[i][0]) else '0',
                        'h': data[i][2] if not pd.isna(data[i][2]) else 0,
                        'w': data[i][4] if not pd.isna(data[i][4]) else 0,
                        'l': data[i][3] if not pd.isna(data[i][3]) else 0,
                        'is_packed': pack(data[i][5])
                    }
                    dop.append(data[i])
            data = dop
            temp_inf = await session.execute(select(Info.id))
            temp_inf = temp_inf.scalars().all()
            cnt = 0
            dop = []
            for row in data:
                # print(row)
                if row['id'] in dop and (row['id'] == 'insp' or row['id'] == 'sh' or row['id'] == 0 or row['id'] == '23362' or row['id'] == '11332' or row['id'] == 'sf1'):
                    # raise HTTPException(status_code=400, detail="В номенклатуре присутствуют повторяющиеся артикулы")
                    continue
                if row['id'] in temp_inf:
                    await session.execute(update(Info).where(Info.id == row['id']).values(**row))
                    dop.append(row['id'])
                else:
                    await session.execute(insert(Info).values(**row))
                    dop.append(row['id'])
            await session.commit()
            async_engine.echo = True
            
    @staticmethod
    async def insert_order(order):
        print('ins_order')
        async with async_session_factory() as session:
            async_engine.echo = False
            order = order[['Артикул', 'Количество', 'Сумма с НДС']]
            order_id = await session.execute(func.count(Order.id))
            order_id = order_id.scalar() + 1
            await session.execute(insert(Order).values({'id': order_id}))
            await session.flush()
            data = order.values.tolist()
            for i in range(len(data)):
                data[i] = {
                    'order_id': order_id,
                    'info_id': data[i][0],
                    'amount': data[i][1],
                    'price': data[i][2]
            } 
            temp_inf = await session.execute(select(Info.id))
            temp_inf = temp_inf.scalars().all()
            # print(temp_inf)
            for row in data:
                # print(row)
                if row['info_id'] in temp_inf:
                    await session.execute(insert(OrdderConstructor).values(row))
            await session.commit()
            async_engine.echo = True
            return order_id
        
    @staticmethod
    async def get_order_num():
        async with async_session_factory() as session:
            order_id = await session.execute(func.count(Order.id))
            order_id = order_id.scalar()
            return order_id

    @staticmethod
    async def calc_time_and_volume(order_id, flag=False):   
        async with async_session_factory() as session:
            query = (
                select(OrdderConstructor).where(OrdderConstructor.order_id == order_id).options(
                    selectinload(OrdderConstructor.info))
                )
            result = await session.execute(query)
            result = result.scalars().all()
            # print(result)
            build = 0
            #TODO объём считается так, а часы: 100л/ч - упаковка т.е. если товар упаковываемый - на него тратится время сверх, 200л/ч - сборка
            #в эксельке выводить товар: описание артикул, количество - габариты, литраж если есть, если нет - красным
            ans = {}
            ans['data'] = {}
            for row in result:
                build += (row.info.h * row.info.w * row.info.l) * row.amount
            build /= 1000
            packing = 0
            for row in result:
                if row.info.is_packed == pack.yes:
                    packing += (row.info.h * row.info.w * row.info.l) * row.amount
            packing /= 1000
            ans['Сборка'] = build
            ans['Упаковка'] = packing
            ans['Сборка время'] = build / 200 * 60
            ans['Упаковка время'] = packing / 100 * 60
            ans['Всего'] = ans['Сборка время'] + ans['Упаковка время']
            if flag:
                for i in range(len(result)):
                    elem = result[i]
                    ans['data'][i] = {
                        'N': i + 1,
                        'Артикул': elem.info.id,
                        'Номенклатура': elem.info.name,
                        'Количество': elem.amount,
                        'Габариты': f'{elem.info.h}х{elem.info.w}х{elem.info.l}',
                        'Литраж': elem.info.h * elem.info.w * elem.info.l / 1000 * elem.amount,
                        'Сборка': elem.info.h * elem.info.w * elem.info.l / 1000,
                        'Упаковка': elem.info.is_packed.value
                    }

                df = pd.DataFrame(ans['data'].values())
                df.loc[0, 'Сборка(л)'] = ans['Сборка']
                df.loc[0, 'Упаковка(л)'] = ans['Упаковка']
                df.loc[0, 'Сборка время'] = ans['Сборка время']
                df.loc[0, 'Упаковка время'] = ans['Упаковка время']
                df.loc[0, 'Всего время'] = ans['Всего']
                BASE_DIR = os.path.dirname(os.path.abspath(__file__))
                SAVE_DIR = os.path.join(BASE_DIR, 'results')
                os.makedirs(SAVE_DIR, exist_ok=True)
                file_path = os.path.join(SAVE_DIR, 'output.xlsx')
                df.to_excel(file_path, index=False)
                wb = load_workbook(file_path)
                ws = wb.active
                red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                pack_col = None
                liter_col = None
                for col_idx, cell in enumerate(ws[1], start=1):
                    if cell.value == "Литраж":
                        liter_col = col_idx
                    if cell.value == "Упаковка":
                        pack_col = col_idx
                        break
                if pack_col:
                    for row in range(2, ws.max_row + 1):
                        val = ws.cell(row=row, column=liter_col).value
                        try:
                            if float(val) == 0:
                                for col in range(1, pack_col + 1):
                                    ws.cell(row=row, column=col).fill = red_fill
                        except:
                            pass
                for col_idx, cell in enumerate(ws[1], start=1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = len(str(cell.value)) + 5
                ws.column_dimensions['C'].width = 60
                
                
                

                wb.save(file_path)
                return file_path
            else:
                return {
                    'Сборка': build,
                    'Упаковка': packing,
                    'Сборка время': build / 200,
                    'Упаковка время': packing / 100,
                    'Всего': ans['Сборка время'] + ans['Упаковка время']
                }
            
                    
                
            
            
            